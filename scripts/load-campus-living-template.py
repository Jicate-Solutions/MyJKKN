"""
Campus Living — xlsx template loader.

Reads a filled campus-living-seed-template.xlsx and inserts rows into
production Supabase in dependency order (Blocks → Rooms → Beds → etc).

Design:
- Imports LABELS / INSTITUTION_UUIDS / SHEETS directly from the generator
  so there's one source of truth for xlsx ↔ DB mapping.
- Translates human labels ("Boys' Hostel") → DB enum values ("boys").
- Resolves College Name → institution_id via INSTITUTION_UUIDS.
- Resolves composite FKs (Block Code → block_id) via in-memory lookup
  tables built as earlier sheets insert.
- Yes/No boolean, day-of-week int, comma-separated arrays handled automatically.
- Dry-run by default — add --apply to actually write.
- Reads prod creds from .env.production.local (same env-parser quirk as
  the PR verification queries from earlier in the session).

Usage:
    python3 scripts/load-campus-living-template.py <xlsx-path>
        → dry-run, prints what would insert

    python3 scripts/load-campus-living-template.py <xlsx-path> --apply
        → actually inserts via Supabase REST

    python3 scripts/load-campus-living-template.py <xlsx-path> --sheets 1,2,3
        → only process specific sheet numbers

    python3 scripts/load-campus-living-template.py <xlsx-path> --apply --stop-on-error
        → halt on first error (useful during initial seed debugging)

Requirements:
    - openpyxl
    - .env.production.local at MyJKKN repo root with NEXT_PUBLIC_SUPABASE_URL
      and SUPABASE_SERVICE_ROLE_KEY
"""

from __future__ import annotations

import argparse
import importlib.util
import json
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[1]
ENV_PATH = REPO_ROOT / ".env.production.local"
GENERATOR_PATH = REPO_ROOT / "scripts/generate-campus-living-template.py"

# The 7 values that exist in the legacy PG enum `hostel_leave_type_enum`
# (shipped 2026-02). The 9 new codes added 2026-04-22 live ONLY in
# the hostel_leave_types master table, not in the enum. When writing
# to hostel_leave_type_config, the enum column must be NULL for new codes
# (requires migration 20260422100000 to have dropped NOT NULL on it).
LEGACY_HOSTEL_LEAVE_TYPE_ENUM_VALUES = {
    "home_visit", "weekend", "vacation", "emergency",
    "medical", "academic", "night_out",
}


# --- Import generator module (no __main__ side effects) ---------------------
def import_generator() -> Any:
    spec = importlib.util.spec_from_file_location("cl_generator", GENERATOR_PATH)
    mod = importlib.util.module_from_spec(spec)  # type: ignore[arg-type]
    spec.loader.exec_module(mod)  # type: ignore[union-attr]
    return mod


# --- Env parser (handles the stray '\n' gotcha from .env.production.local) --
def load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    with open(ENV_PATH) as f:
        for line in f:
            line = line.rstrip("\n\r")
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            env[k] = v.strip().strip('"').strip("'").rstrip("\\n").rstrip("\\")
    return env


# --- Supabase REST helpers --------------------------------------------------
class Supabase:
    def __init__(self, url: str, key: str, dry_run: bool):
        self.url = url
        self.key = key
        self.dry_run = dry_run
        self.session_headers = {
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        }

    def insert(self, table: str, row: dict[str, Any]) -> dict[str, Any]:
        """POST /rest/v1/<table> with a single row. Returns inserted row dict."""
        if self.dry_run:
            return {"_dry_run": True, "_table": table, **row}
        body = json.dumps([row]).encode()
        req = urllib.request.Request(
            f"{self.url}/rest/v1/{table}",
            data=body,
            method="POST",
            headers=self.session_headers,
        )
        try:
            with urllib.request.urlopen(req) as r:
                data = json.loads(r.read() or "[]")
                return data[0] if data else {}
        except urllib.error.HTTPError as e:
            err_body = e.read().decode()
            raise RuntimeError(f"{e.code} {err_body[:300]}") from None

    def query(self, path: str) -> list[dict[str, Any]]:
        """GET /rest/v1/<path>. Returns list of rows."""
        req = urllib.request.Request(
            f"{self.url}/rest/v1/{path}",
            headers={"apikey": self.key, "Authorization": f"Bearer {self.key}"},
        )
        with urllib.request.urlopen(req) as r:
            return json.loads(r.read() or "[]")


# --- Value translation ------------------------------------------------------
def build_label_maps(LABELS: dict) -> dict[str, dict[str, str]]:
    """For each enum group, build a label → db_value map for reverse lookup."""
    return {
        group: {label: db_val for label, db_val in pairs}
        for group, pairs in LABELS.items()
    }


def translate_value(
    value: Any,
    validation_key: str | None,
    label_maps: dict[str, dict[str, str]],
    db_column: str,
) -> Any:
    """Translate a raw cell value to the DB-appropriate type/value.

    - Institution dropdown: passthrough (loader handles separately)
    - yes_no: "Yes" → True, "No" → False
    - day_of_week: "Monday" → 1
    - enum labels: "Boys' Hostel" → "boys"
    - Comma-separated arrays (items, collection_days, etc.) → list of strings
    - None / empty string → None
    """
    if value is None or value == "":
        return None

    # Strip strings
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None

    # Booleans
    if validation_key == "yes_no":
        return label_maps["yes_no"].get(str(value)) == "TRUE"

    # Day of week → integer
    if validation_key == "day_of_week":
        mapped = label_maps["day_of_week"].get(str(value))
        return int(mapped) if mapped is not None else value

    # Enum (hostel_type, room_type, etc.)
    if validation_key in label_maps:
        return label_maps[validation_key].get(str(value), value)

    # Array-type columns (comma-separated)
    array_cols = {
        "items", "special_items", "dietary_tags",
        "collection_days", "delivery_days",
        "assigned_floors",
    }
    if db_column in array_cols and isinstance(value, str):
        return [v.strip() for v in value.split(",") if v.strip()]

    return value


# --- Row extraction ---------------------------------------------------------
YELLOW_FILL_HEX = "00FEF3C7"


def extract_data_rows(ws, columns: list[tuple]) -> list[dict[str, Any]]:
    """Yield data rows from a sheet, skipping header + hint + example (yellow)
    rows, stopping at first all-empty row.

    Template layout:
        Row 1 — description
        Row 2 — headers
        Row 3 — hints
        Row 4+ — example rows (yellow fill) OR user data (green fill) OR empty
    """
    rows = []
    for r in range(4, ws.max_row + 1):
        # Skip if first cell has yellow fill (example row)
        first_cell = ws.cell(row=r, column=1)
        if first_cell.fill and first_cell.fill.fgColor and first_cell.fill.fgColor.rgb == YELLOW_FILL_HEX:
            continue

        # Collect values
        raw = {}
        any_value = False
        for ci, col_spec in enumerate(columns, 1):
            label, _required, _hint, validation_key, db_column = col_spec
            cell_val = ws.cell(row=r, column=ci).value
            raw[db_column] = {
                "value": cell_val,
                "label": label,
                "validation_key": validation_key,
            }
            if cell_val not in (None, ""):
                any_value = True

        # All-empty → end of data
        if not any_value:
            break

        rows.append(raw)
    return rows


# --- Per-sheet processors ---------------------------------------------------
class LoaderContext:
    """Accumulates state across sheets (inserted IDs for FK resolution)."""

    def __init__(self, sb: Supabase, institutions: dict[str, str]):
        self.sb = sb
        self.institutions = institutions  # name → UUID
        # FK caches: lookup key → UUID
        self.blocks: dict[tuple, str] = {}    # (college_uuid, block_code) → block_id
        self.rooms: dict[tuple, str] = {}     # (college_uuid, block_code, room_number) → room_id
        self.caterers: dict[tuple, str] = {}  # (college_uuid, caterer_name) → caterer_id
        self.academic_years: dict[tuple, str] = {}  # (college_uuid, label) → academic_year_id

    def resolve_institution(self, name: str) -> str:
        uuid = self.institutions.get(name)
        if not uuid:
            raise ValueError(
                f"College '{name}' not in REF Institutions. Check exact spelling."
            )
        return uuid

    def resolve_academic_year(self, college_uuid: str, label: str) -> str:
        key = (college_uuid, label)
        if key in self.academic_years:
            return self.academic_years[key]
        # Query prod
        path = (
            f"academic_years?institution_id=eq.{college_uuid}"
            f"&year_name=eq.{urllib.parse.quote(label)}&select=id"
        )
        rows = self.sb.query(path)
        if not rows:
            raise ValueError(
                f"Academic year '{label}' not found for institution {college_uuid}. "
                "Create it in Organization → Academic Years first."
            )
        self.academic_years[key] = rows[0]["id"]
        return rows[0]["id"]


def process_sheet(
    ctx: LoaderContext,
    sheet_name: str,
    columns: list[tuple],
    raw_rows: list[dict],
    label_maps: dict,
) -> tuple[int, list[str]]:
    """Insert rows from one sheet. Returns (inserted_count, errors)."""
    table_map = {
        "1. Hostel Blocks":           "hostel_blocks",
        "2. Hostel Rooms":            "hostel_rooms",
        "3. Hostel Beds":             "hostel_beds",
        "4. Hostel Wardens":          "hostel_wardens",
        "5. Fee Config":              "hostel_fee_config",
        "6. Leave Types":             "hostel_leave_type_config",
        "7. Mess Caterers":           "mess_caterers",
        "8. Mess Menus":              "mess_menus",
        "9. Maintenance SLA":         "hostel_maintenance_sla_config",
        "10. Cleaning Schedules":     "hostel_cleaning_schedules",
        "11. Laundry Config":         "hostel_laundry_configs",
        "12. Safety Equipment":       "hostel_safety_equipment",
    }
    table = table_map.get(sheet_name)
    if not table:
        return (0, [f"No table mapping for sheet '{sheet_name}'"])

    inserted = 0
    errors: list[str] = []

    for idx, raw in enumerate(raw_rows, 1):
        try:
            row = translate_row(ctx, sheet_name, columns, raw, label_maps)
            if row is None:
                continue  # row is flagged as skipped
            result = ctx.sb.insert(table, row)
            inserted += 1
            cache_fk_lookups(ctx, sheet_name, raw, result)
        except Exception as e:
            errors.append(f"  row {idx}: {e}")

    return inserted, errors


def translate_row(
    ctx: LoaderContext,
    sheet_name: str,
    columns: list[tuple],
    raw: dict,
    label_maps: dict,
) -> dict[str, Any] | None:
    """Apply all transformations and FK resolution to produce a DB-ready row."""
    row: dict[str, Any] = {}

    for col_spec in columns:
        _label, required, _hint, validation_key, db_column = col_spec

        # Underscore-prefixed columns are "virtual" — handled separately
        if db_column.startswith("_"):
            continue

        raw_info = raw.get(db_column)
        if raw_info is None:
            continue

        value = translate_value(
            raw_info["value"], validation_key, label_maps, db_column
        )

        # Institution resolution
        if validation_key == "institution":
            if not value:
                if required:
                    raise ValueError(f"College Name is required")
                continue
            value = ctx.resolve_institution(str(value))

        # Skip None if field is optional
        if value is None:
            if required:
                raise ValueError(f"'{_label}' is required but empty")
            continue

        row[db_column] = value

    # Apply sheet-specific FK resolutions
    row = apply_fk_resolutions(ctx, sheet_name, columns, raw, row)

    return row


def apply_fk_resolutions(
    ctx: LoaderContext,
    sheet_name: str,
    columns: list[tuple],
    raw: dict,
    row: dict,
) -> dict:
    """Resolve virtual columns (_block_code, _room_number, _caterer_name,
    _academic_year_label, _staff_name, _employee_id) to real FK UUIDs."""

    def get_virtual(key: str) -> Any:
        info = raw.get(key)
        return info["value"].strip() if info and isinstance(info.get("value"), str) else (info["value"] if info else None)

    # Rooms: block_id from (institution, block_code)
    if sheet_name == "2. Hostel Rooms":
        block_code = get_virtual("_block_code")
        if block_code:
            key = (row.get("institution_id"), block_code)
            bid = ctx.blocks.get(key)
            if not bid:
                raise ValueError(f"Block Code '{block_code}' not found (insert Blocks sheet first)")
            row["block_id"] = bid

    # Beds: room_id from (institution, block_code, room_number)
    if sheet_name == "3. Hostel Beds":
        block_code = get_virtual("_block_code")
        room_number = get_virtual("_room_number")
        if block_code and room_number:
            key = (row.get("institution_id"), block_code, str(room_number))
            rid = ctx.rooms.get(key)
            if not rid:
                raise ValueError(
                    f"Room '{block_code}/{room_number}' not found (insert Rooms sheet first)"
                )
            row["room_id"] = rid

    # Wardens: block_id (optional), + staff_id/user_id resolution via Employee ID
    if sheet_name == "4. Hostel Wardens":
        block_code = get_virtual("_block_code")
        if block_code:
            key = (row.get("institution_id"), block_code)
            bid = ctx.blocks.get(key)
            if bid:
                row["block_id"] = bid

        emp_id = get_virtual("_employee_id")
        if emp_id:
            # Query staff by employee_id
            staff_rows = ctx.sb.query(
                f"staff?employee_id=eq.{urllib.parse.quote(str(emp_id))}"
                "&select=id,user_id&limit=1"
            )
            if not staff_rows:
                raise ValueError(f"Staff with Employee ID '{emp_id}' not found")
            row["staff_id"] = staff_rows[0]["id"]
            row["user_id"] = staff_rows[0].get("user_id") or staff_rows[0]["id"]

    # Fee Config: academic_year_id from label
    if sheet_name == "5. Fee Config":
        year_label = get_virtual("_academic_year_label")
        if year_label:
            row["academic_year_id"] = ctx.resolve_academic_year(
                row.get("institution_id"), str(year_label)
            )

    # Leave Types: resolve leave_type_id from the master (hostel_leave_types);
    # null out the legacy enum column for new codes that aren't in the
    # 7-value hostel_leave_type_enum. Requires migration
    # 20260422100000_hostel_leave_type_config_nullable_enum.sql applied first.
    if sheet_name == "6. Leave Types":
        institution_id = row.get("institution_id")
        type_code = row.get("leave_type")  # already translated to enum-string via LABELS
        if institution_id and type_code:
            path = (
                f"hostel_leave_types?institution_id=eq.{institution_id}"
                f"&leave_type_code=eq.{urllib.parse.quote(str(type_code))}"
                f"&select=id&limit=1"
            )
            master_rows = [] if ctx.sb.dry_run else ctx.sb.query(path)
            if ctx.sb.dry_run:
                row["leave_type_id"] = f"dry-hlt-{type_code}"
            elif not master_rows:
                raise ValueError(
                    f"Leave type '{type_code}' not found in hostel_leave_types "
                    f"master for this institution. Seed the master via migration "
                    f"first (it should already contain the 16 defaults)."
                )
            else:
                row["leave_type_id"] = master_rows[0]["id"]
            # Null the legacy enum column for new (post-2026-04-22) codes so
            # the PG enum constraint doesn't reject the insert.
            if type_code not in LEGACY_HOSTEL_LEAVE_TYPE_ENUM_VALUES:
                row.pop("leave_type", None)

    # Mess Menus: caterer_id from caterer_name, optional block_id
    if sheet_name == "8. Mess Menus":
        caterer_name = get_virtual("_caterer_name")
        if caterer_name:
            key = (row.get("institution_id"), caterer_name)
            cid = ctx.caterers.get(key)
            if not cid:
                raise ValueError(
                    f"Caterer '{caterer_name}' not found (insert Mess Caterers sheet first)"
                )
            row["caterer_id"] = cid
        block_code = get_virtual("_block_code")
        if block_code:
            key = (row.get("institution_id"), block_code)
            row["block_id"] = ctx.blocks.get(key)  # optional, can be None

    # Cleaning/Laundry/Safety: optional block_id from block_code
    if sheet_name in ("10. Cleaning Schedules", "11. Laundry Config", "12. Safety Equipment"):
        block_code = get_virtual("_block_code")
        if block_code:
            key = (row.get("institution_id"), block_code)
            row["block_id"] = ctx.blocks.get(key)

    return row


def cache_fk_lookups(ctx: LoaderContext, sheet_name: str, raw: dict, result: dict):
    """After successful insert, cache the new UUID so downstream sheets can
    resolve FKs by human key."""
    if ctx.sb.dry_run:
        # Fake UUID for dry-run lookups (won't hit DB but test sheet dependencies)
        result = {**result, "id": f"dry-{result.get('_table', '')}-{len(ctx.blocks) + len(ctx.rooms)}"}

    if sheet_name == "1. Hostel Blocks":
        code = raw.get("code", {}).get("value")
        inst = ctx.institutions.get(raw.get("institution_id", {}).get("value", ""))
        if code and inst:
            ctx.blocks[(inst, code.strip())] = result["id"]
    elif sheet_name == "2. Hostel Rooms":
        block_code = raw.get("_block_code", {}).get("value")
        room_number = raw.get("room_number", {}).get("value")
        inst = ctx.institutions.get(raw.get("institution_id", {}).get("value", ""))
        if block_code and room_number and inst:
            ctx.rooms[(inst, block_code.strip(), str(room_number).strip())] = result["id"]
    elif sheet_name == "7. Mess Caterers":
        name = raw.get("name", {}).get("value")
        inst = ctx.institutions.get(raw.get("institution_id", {}).get("value", ""))
        if name and inst:
            ctx.caterers[(inst, name.strip())] = result["id"]


# --- Main -------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser(description="Load campus-living seed xlsx into prod")
    ap.add_argument("xlsx", type=Path, help="path to filled xlsx template")
    ap.add_argument("--apply", action="store_true",
                    help="actually insert rows (default: dry-run)")
    ap.add_argument("--sheets", type=str, default=None,
                    help="comma-sep sheet numbers, e.g. '1,2,3'. Default: all")
    ap.add_argument("--stop-on-error", action="store_true",
                    help="halt on first row error")
    args = ap.parse_args()

    if not args.xlsx.exists():
        print(f"ERROR: {args.xlsx} not found")
        return 2

    # Defer openpyxl import so --help works without it
    from openpyxl import load_workbook

    gen = import_generator()
    env = load_env()
    sb = Supabase(env["NEXT_PUBLIC_SUPABASE_URL"], env["SUPABASE_SERVICE_ROLE_KEY"],
                  dry_run=not args.apply)
    ctx = LoaderContext(sb, gen.INSTITUTION_UUIDS)
    label_maps = build_label_maps(gen.LABELS)

    wb = load_workbook(args.xlsx, data_only=True)

    sheets_filter = None
    if args.sheets:
        sheets_filter = set(s.strip() for s in args.sheets.split(","))

    mode = "DRY-RUN" if not args.apply else "APPLY"
    print(f"=== Campus Living Loader — {mode} ===")
    print(f"File: {args.xlsx}")
    print(f"Target: {env['NEXT_PUBLIC_SUPABASE_URL']}")
    print()

    total_inserted = 0
    total_errors = 0

    for sheet_name, _desc, columns, _samples in gen.SHEETS:
        # e.g. sheet_name starts with "1. ", "2. ", etc.
        m = re.match(r"^(\d+)\.", sheet_name)
        sheet_num = m.group(1) if m else ""
        if sheets_filter and sheet_num not in sheets_filter:
            continue

        ws = wb[sheet_name]
        raw_rows = extract_data_rows(ws, columns)
        if not raw_rows:
            print(f"[{sheet_name}] no data rows — skipping")
            continue

        print(f"[{sheet_name}] {len(raw_rows)} rows to process...")
        inserted, errors = process_sheet(ctx, sheet_name, columns, raw_rows, label_maps)
        total_inserted += inserted
        total_errors += len(errors)

        status = "✓" if not errors else "✗"
        print(f"  {status} inserted={inserted}, errors={len(errors)}")
        for err in errors:
            print(err)

        if errors and args.stop_on_error:
            print("STOP-ON-ERROR triggered.")
            break

    print()
    print(f"=== Summary ===")
    print(f"Inserted: {total_inserted}")
    print(f"Errors:   {total_errors}")
    print(f"Mode:     {mode}")

    return 0 if total_errors == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
