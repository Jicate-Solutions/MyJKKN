"""
Parse the JKKN Dental hostel inventory Excel and emit a normalized JSON file.

Input:  /Users/omm/Downloads/vnd.openxmlformats-officedocument.spreadsheetml.sheet&rendition=1.xlsx
Output: scripts/hostel-import/parsed.json
"""
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

SRC = '/Users/omm/Downloads/vnd.openxmlformats-officedocument.spreadsheetml.sheet&rendition=1.xlsx'
OUT = Path(__file__).parent / 'parsed.json'

FLOOR_MAP = {
    'GROUND': 0,
    'FIRST FLOOR': 1,
    'SECOND FLOOR': 2,
    '2ND FLR': 2,
    '3RD FLR': 3,
}

def norm(s):
    if s is None:
        return None
    return str(s).strip()

def norm_status(s):
    s = norm(s)
    if not s:
        return None
    s = re.sub(r'\s+', ' ', s)
    return s.upper()

def parse_dept(dept):
    """Return (program, year) e.g. ('BDS', 2) or ('PHARMD', 1)."""
    if not dept:
        return (None, None)
    d = re.sub(r'\s+', ' ', dept.upper().strip())
    # Normalize PHARM D / PHARMD
    d = d.replace('PHARM D', 'PHARMD')
    m = re.match(r'(BDS|PHARMD)\s*(\d+)(?:ST|ND|RD|TH)?\s*(?:YR|YEAR)', d)
    if m:
        return (m.group(1), int(m.group(2)))
    return (None, None)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    out = {
        'blocks': [
            {'code': 'A', 'name': 'A Block', 'hostel_type': 'boys', 'total_floors': 3},
            {'code': 'B', 'name': 'B Block', 'hostel_type': 'boys', 'total_floors': 3},
            {'code': 'C', 'name': 'C Block', 'hostel_type': 'girls', 'total_floors': 4},
        ],
        'rooms': [],         # list of {block_code, room_number, floor, capacity, status, tier_key, ac, tiles}
        'residents': [],     # list of {block_code, room_number, raw_name, normalized_name, dept_program, dept_year}
    }

    # A BLOCK + B BLOCK have same shape
    for sheet_name, block_code in [('A BLOCK', 'A'), ('B BLOCK', 'B')]:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            room_num, floor_raw, capacity, status_raw, tier_raw = row[:5]
            if room_num is None:
                continue
            floor = FLOOR_MAP.get(norm(floor_raw))
            if floor is None:
                print(f'WARN: {sheet_name} room {room_num} unknown floor {floor_raw}', file=sys.stderr)
                continue
            status_n = norm_status(status_raw)
            tier_n = norm(tier_raw)
            tier_key = 'premium' if tier_n == 'PREMIUM' else 'standard'
            out['rooms'].append({
                'block_code': block_code,
                'room_number': str(room_num),
                'floor': floor,
                'capacity': int(capacity) if isinstance(capacity, (int, float)) else 1,
                'raw_status': status_n,
                'tier_key': tier_key,
                'ac': False,
                'tiles': False,
            })

    # C BLOCK
    ws = wb['C BLOCK']
    # Rooms are deduplicated (multiple resident rows share a room)
    seen_rooms = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        room_num, floor_raw, name_raw, dept_raw, capacity, status_raw, ac_raw, tier_raw, tiles_raw = row[:9]
        if room_num is None:
            continue
        floor = FLOOR_MAP.get(norm(floor_raw))
        if floor is None:
            print(f'WARN: C BLOCK room {room_num} unknown floor {floor_raw}', file=sys.stderr)
            continue
        status_n = norm_status(status_raw)
        tier_n = norm(tier_raw)
        tier_key = 'premium' if tier_n == 'PREMIUM' else 'standard'
        ac_n = norm(ac_raw)
        ac = ac_n is not None and ac_n.upper().strip() == 'AC'
        tiles_n = norm(tiles_raw)
        tiles = tiles_n == 'TILES'
        # Capacity may be only on first occurrence — propagate
        key = (room_num, floor)
        if key not in seen_rooms:
            seen_rooms[key] = {
                'block_code': 'C',
                'room_number': str(room_num),
                'floor': floor,
                'capacity': int(capacity) if isinstance(capacity, (int, float)) else 1,
                'raw_status': status_n,
                'tier_key': tier_key,
                'ac': ac,
                'tiles': tiles,
            }
        else:
            # Fill capacity if it was None before
            existing = seen_rooms[key]
            if existing['capacity'] in (None, 1) and isinstance(capacity, (int, float)):
                existing['capacity'] = int(capacity)
            # AC/tiles can be set if any row has it
            if ac:
                existing['ac'] = True
            if tiles:
                existing['tiles'] = True

        # Resident row?
        if name_raw and isinstance(name_raw, str) and name_raw.strip():
            program, year = parse_dept(dept_raw)
            out['residents'].append({
                'block_code': 'C',
                'room_number': str(room_num),
                'floor': floor,
                'raw_name': name_raw.strip(),
                'normalized_name': normalize_name(name_raw),
                'raw_dept': norm(dept_raw),
                'program': program,
                'year': year,
            })

    out['rooms'].extend(seen_rooms.values())

    OUT.write_text(json.dumps(out, indent=2))
    print(f'Wrote {OUT}')
    print(f'Blocks: {len(out["blocks"])}')
    print(f'Rooms: {len(out["rooms"])} (A={sum(1 for r in out["rooms"] if r["block_code"]=="A")} B={sum(1 for r in out["rooms"] if r["block_code"]=="B")} C={sum(1 for r in out["rooms"] if r["block_code"]=="C")})')
    print(f'Residents (C BLOCK named girls): {len(out["residents"])}')
    # Program breakdown
    prog = Counter((r['program'], r['year']) for r in out['residents'])
    print('Programs:', dict(prog))
    # Capacity check
    print('A capacities:', dict(Counter(r['capacity'] for r in out['rooms'] if r['block_code']=='A')))
    print('B capacities:', dict(Counter(r['capacity'] for r in out['rooms'] if r['block_code']=='B')))
    print('C capacities:', dict(Counter(r['capacity'] for r in out['rooms'] if r['block_code']=='C')))


def normalize_name(name):
    """Lowercase, strip dots, collapse spaces, keep alpha tokens only."""
    if not name:
        return ''
    n = name.lower()
    n = re.sub(r'[^a-z\s]', ' ', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n


if __name__ == '__main__':
    main()
