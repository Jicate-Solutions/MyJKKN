"""
Generate Campus Living bulk-seed Excel template — WARDEN-FRIENDLY VERSION.

Design principles (v2):
- ZERO UUIDs visible to the user
- Every dropdown uses plain English labels, NOT database enum codes
- Every boolean dropdown is Yes/No
- Column headers are Title Case English ("Block Name" not "name", "Type of Hostel" not "hostel_type")
- Wardens identify things by human keys: College Name, Block Code, Room Number, Employee ID, Register Number
- A companion import loader (scripts/load-campus-living-template.py) does ALL translation:
  * College Name → institution_id UUID
  * "Boys' Hostel" → "boys" enum value
  * "Yes" → TRUE boolean
  * (College + Block Code) → block_id UUID
  * (College + Block Code + Room Number) → room_id UUID
  * Employee ID → staff_id/user_id UUID
  * Register Number → learner_id UUID
  * Academic year label "2026-27" → academic_year_id UUID

Usage:
    python3 scripts/generate-campus-living-template.py
Output:
    docs/templates/campus-living/campus-living-seed-template.xlsx
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

OUT = Path(__file__).resolve().parents[1] / "docs/templates/campus-living/campus-living-seed-template.xlsx"

# --- Styles ----------------------------------------------------------------
BLUE = PatternFill("solid", fgColor="1E40AF")
RED = PatternFill("solid", fgColor="DC2626")
GRAY_HEAD = PatternFill("solid", fgColor="374151")
GRAY_LIGHT = PatternFill("solid", fgColor="F3F4F6")
YELLOW = PatternFill("solid", fgColor="FEF3C7")
GREEN = PatternFill("solid", fgColor="DCFCE7")
ORANGE = PatternFill("solid", fgColor="FB923C")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
HINT_FONT = Font(color="6B7280", italic=True, size=9)
TITLE_FONT = Font(color="1E3A8A", bold=True, size=14)
BORDER = Border(left=Side(style="thin", color="D1D5DB"),
                right=Side(style="thin", color="D1D5DB"),
                top=Side(style="thin", color="D1D5DB"),
                bottom=Side(style="thin", color="D1D5DB"))

# --- Institutions (pulled 2026-04-20 from prod) -----------------------------
INSTITUTIONS = [
    "JKKN College of Engineering and Technology",
    "JKKN College of Pharmacy",
    "JKKN Dental College and Hospital",
    "JKKN College of Allied Health Sciences",
    "JKKN College of Nursing and Research",
    "JKKN College of Arts and Science (Aided)",
    "JKKN College of Arts and Science (Self)",
    "JKKN College of Education",
]

# Kept for reference sheet only — loader maps name → UUID, user never sees UUIDs
INSTITUTION_UUIDS = {
    "JKKN College of Engineering and Technology": "5de4fba1-4564-41ed-8c73-5d948b74b843",
    "JKKN College of Pharmacy": "5736d86f-5dab-4b7f-9aa1-b3bb1a2dd334",
    "JKKN Dental College and Hospital": "e8fbe8aa-c44e-41aa-a44b-39dab2c8b9a5",
    "JKKN College of Allied Health Sciences": "9c1554e8-12a2-4b76-a9d6-8242bb05eba1",
    "JKKN College of Nursing and Research": "70e54e51-9b98-4e07-9534-a85310609bfd",
    "JKKN College of Arts and Science (Aided)": "a33138b6-4eea-4675-941f-1071bf88b127",
    "JKKN College of Arts and Science (Self)": "b0b8a724-7c65-4f07-8047-2a38e8100ad5",
    "JKKN College of Education": "9380358f-7020-4c23-89c3-e9538b47cf33",
}

# --- Label → DB value maps -------------------------------------------------
# Each list is ordered for the dropdown. Labels are shown; loader converts to the DB value.
LABELS = {
    "hostel_type": [
        ("Boys' Hostel", "boys"),
        ("Girls' Hostel", "girls"),
        ("Mixed (Boys + Girls)", "mixed"),
        ("Staff Quarters", "staff"),
        ("International Students", "international"),
        ("Married Students", "married"),
        ("Working Women", "working_women"),
        ("Medical Students", "medical"),
    ],
    "block_status": [
        ("Active", "active"),
        ("Under Maintenance", "under_maintenance"),
        ("Closed", "closed"),
    ],
    "room_type": [
        ("Single (1 bed)", "single"),
        ("Double (2 beds)", "double"),
        ("Triple (3 beds)", "triple"),
        ("Quad (4 beds)", "quad"),
        ("Dormitory (5+ beds)", "dormitory"),
    ],
    "ac_status": [
        ("AC", "ac"),
        ("Non-AC", "non_ac"),
        ("Cooler", "cooler"),
    ],
    "bed_type": [
        ("Single Bed", "single"),
        ("Bunk Bed — Upper", "bunk_upper"),
        ("Bunk Bed — Lower", "bunk_lower"),
    ],
    "bed_status": [
        ("Available", "available"),
        ("Occupied", "occupied"),
        ("Reserved", "reserved"),
        ("Under Maintenance", "maintenance"),
    ],
    "room_status": [
        ("Available", "available"),
        ("Partially Occupied", "partially_occupied"),
        ("Full", "full"),
        ("Under Maintenance", "maintenance"),
        ("Reserved", "reserved"),
        ("Closed", "closed"),
    ],
    "warden_designation": [
        ("Chief Warden", "chief_warden"),
        ("Warden", "warden"),
        ("Deputy Warden", "deputy_warden"),
        ("Floor Supervisor", "floor_supervisor"),
        ("Night Watcher", "night_watcher"),
    ],
    "warden_shift": [
        ("Day Shift", "day"),
        ("Night Shift", "night"),
        ("Full-Time (Day + Night)", "full_time"),
    ],
    "allocation_type": [
        ("Fresh (new admission)", "fresh"),
        ("Renewal (returning student)", "renewal"),
        ("Transfer (between blocks/rooms)", "transfer"),
        ("Temporary (short stay)", "temporary"),
    ],
    "food_preference": [
        ("Vegetarian", "vegetarian"),
        ("Non-Vegetarian", "non_vegetarian"),
        ("Vegan", "vegan"),
        ("Jain", "jain"),
        ("Eggetarian (veg + egg)", "eggetarian"),
    ],
    "electricity_charges": [
        ("Included in hostel fees", "included"),
        ("Metered (based on actual usage)", "metered"),
        ("Fixed monthly charge", "fixed_monthly"),
    ],
    "leave_type": [
        ("Home Visit", "home_visit"),
        ("Weekend Out", "weekend"),
        ("Vacation", "vacation"),
        ("Emergency", "emergency"),
        ("Medical", "medical"),
        ("Academic (exam / conference)", "academic"),
        ("Night Out", "night_out"),
    ],
    "cleaning_type": [
        ("Daily Sweeping", "daily_sweep"),
        ("Daily Mopping", "daily_mop"),
        ("Toilet Cleaning", "toilet_cleaning"),
        ("Common Area Cleaning", "common_area"),
        ("Deep Cleaning", "deep_cleaning"),
        ("Window Cleaning", "window_cleaning"),
        ("Water Tank Cleaning", "water_tank"),
        ("Disinfection / Pest Control", "disinfection"),
        ("Other", "other"),
    ],
    "cleaning_frequency": [
        ("Daily", "daily"),
        ("Weekly", "weekly"),
        ("Every 2 Weeks", "biweekly"),
        ("Monthly", "monthly"),
        ("Every 3 Months", "quarterly"),
        ("Every 6 Months", "half_yearly"),
        ("Yearly", "yearly"),
    ],
    "laundry_service_type": [
        ("In-House (hostel staff)", "in_house"),
        ("External Vendor", "vendor"),
    ],
    "caterer_status": [
        ("Active", "active"),
        ("Contract Ended", "contract_ended"),
        ("Suspended", "suspended"),
        ("Blacklisted", "blacklisted"),
    ],
    "billing_model": [
        ("Fixed Monthly Fee", "fixed_monthly"),
        ("Per-Meal Charging", "per_meal"),
        ("BDMR (Booked but Did not Eat)", "bdmr"),
        ("Semester Advance", "semester_advance"),
    ],
    "meal_type": [
        ("Breakfast", "breakfast"),
        ("Lunch", "lunch"),
        ("Snacks / Tea", "snacks"),
        ("Dinner", "dinner"),
    ],
    "menu_status": [
        ("Planned", "planned"),
        ("Confirmed", "confirmed"),
        ("Served", "served"),
        ("Cancelled", "cancelled"),
    ],
    "maintenance_category": [
        ("Electrical", "electrical"),
        ("Plumbing", "plumbing"),
        ("Civil / Structural", "civil"),
        ("Pest Control", "pest_control"),
        ("Cleaning", "cleaning"),
        ("Internet / WiFi", "internet"),
        ("Water Supply", "water_supply"),
        ("Furniture", "furniture"),
        ("Safety Equipment", "safety"),
        ("Other", "other"),
    ],
    "maintenance_priority": [
        ("Critical (fix immediately)", "critical"),
        ("High (same day)", "high"),
        ("Medium (within 2-3 days)", "medium"),
        ("Low (within a week)", "low"),
    ],
    "yes_no": [
        ("Yes", "TRUE"),
        ("No", "FALSE"),
    ],
    "day_of_week": [
        ("Sunday", "0"),
        ("Monday", "1"),
        ("Tuesday", "2"),
        ("Wednesday", "3"),
        ("Thursday", "4"),
        ("Friday", "5"),
        ("Saturday", "6"),
    ],
}


# --- Sheet definitions -----------------------------------------------------
# Each sheet: (display_name, description, columns, example_rows)
# Column spec: (header_label, required_bool, hint, validation_key_or_None, db_column_name)
# validation_key values: None | enum name (in LABELS) | "yes_no" | "institution" | "day_of_week"

INST = "institution"

SHEETS = [
    ("1. Hostel Blocks",
     "Start here. One row per physical hostel block (e.g. Boys' Block A, Girls' Block B).",
     [
        ("College Name", True, "Pick your college from the dropdown", INST, "institution_id"),
        ("Block Name", True, "e.g. 'Boys Block A', 'Anna Block'", None, "name"),
        ("Block Code", True, "Short unique code e.g. 'BLK-A', 'AB'. This is how Rooms sheet links back.", None, "code"),
        ("Type of Hostel", True, "Pick from dropdown", "hostel_type", "hostel_type"),
        ("Total Floors", True, "How many floors? e.g. 4 (ground=1, first=2...)", None, "total_floors"),
        ("Address", False, "Physical address of the block", None, "address"),
        ("Contact Phone", False, "10-digit warden desk phone", None, "contact_phone"),
        ("Weekday Curfew Time", False, "Lock-in time Mon-Fri (HH:MM, e.g. 22:00)", None, "curfew_time_weekday"),
        ("Weekend Curfew Time", False, "Lock-in time Sat-Sun (HH:MM, e.g. 23:00)", None, "curfew_time_weekend"),
        ("Visiting Hours Start", False, "Visitor entry start time (HH:MM)", None, "visiting_hours_start"),
        ("Visiting Hours End", False, "Visitor entry end time (HH:MM)", None, "visiting_hours_end"),
        ("Status", False, "Pick from dropdown (default: Active)", "block_status", "status"),
     ],
     [
        ["JKKN College of Engineering and Technology", "Boys Block A", "BLK-A", "Boys' Hostel", 4,
         "Main campus, Gate 1", "9876543210", "22:00", "23:00", "16:00", "19:00", "Active"],
        ["JKKN College of Engineering and Technology", "Girls Block A", "GRL-A", "Girls' Hostel", 3,
         "Main campus, Gate 2", "9876543211", "21:30", "22:30", "16:00", "18:30", "Active"],
     ]),

    ("2. Hostel Rooms",
     "Fill this AFTER Blocks sheet is loaded. Reference each block by its Block Code from Blocks sheet.",
     [
        ("College Name", True, "Same college as the block", INST, "institution_id"),
        ("Block Code", True, "Must match a Block Code from the Blocks sheet exactly", None, "_block_code"),
        ("Room Number", True, "e.g. '101', 'A-203'", None, "room_number"),
        ("Floor", True, "Integer: 1 = ground, 2 = first, etc.", None, "floor"),
        ("Room Type", True, "Pick from dropdown", "room_type", "room_type"),
        ("AC Status", True, "Pick from dropdown", "ac_status", "ac_status"),
        ("Capacity", True, "How many beds can fit? (matches Room Type)", None, "capacity"),
        ("Wheelchair Accessible", False, "Yes / No", "yes_no", "is_accessible"),
        ("Attached Bathroom", False, "Yes / No", "yes_no", "has_attached_bathroom"),
        ("Annual Fee (₹)", False, "Override of fee config for this specific room (optional)", None, "annual_fee"),
        ("Status", False, "Pick from dropdown (default: Available)", "room_status", "status"),
        ("Notes", False, "Any notes about maintenance, layout, etc.", None, "maintenance_notes"),
     ],
     [
        ["JKKN College of Engineering and Technology", "BLK-A", "101", 1, "Double (2 beds)", "Non-AC", 2, "No", "Yes", "", "Available", ""],
        ["JKKN College of Engineering and Technology", "BLK-A", "102", 1, "Triple (3 beds)", "AC", 3, "No", "Yes", "", "Available", ""],
     ]),

    ("3. Hostel Beds",
     "Individual beds inside each room. Link to room by College + Block Code + Room Number.",
     [
        ("College Name", True, "Same as the room", INST, "institution_id"),
        ("Block Code", True, "From Blocks sheet", None, "_block_code"),
        ("Room Number", True, "From Rooms sheet", None, "_room_number"),
        ("Bed Number/Label", True, "e.g. 'A', 'B', '1', '2'", None, "bed_number"),
        ("Bed Type", True, "Pick from dropdown", "bed_type", "bed_type"),
        ("Status", False, "Pick from dropdown (default: Available)", "bed_status", "status"),
     ],
     [
        ["JKKN College of Engineering and Technology", "BLK-A", "101", "A", "Single Bed", "Available"],
        ["JKKN College of Engineering and Technology", "BLK-A", "101", "B", "Single Bed", "Available"],
     ]),

    ("4. Hostel Wardens",
     "Warden/supervisor staff assignments. Identify staff by Employee ID (loader looks up the user).",
     [
        ("College Name", True, "Warden's college", INST, "institution_id"),
        ("Staff Name", True, "Warden's full name (for your reference)", None, "_staff_name"),
        ("Employee ID", True, "Staff employee ID from HR system. Loader uses this to find the user.", None, "_employee_id"),
        ("Block Code", False, "Which block they manage (leave blank for institution-wide role)", None, "_block_code"),
        ("Designation", True, "Pick from dropdown", "warden_designation", "designation"),
        ("Phone", True, "10-digit mobile", None, "phone"),
        ("Lives in Hostel", False, "Is the warden residential? Yes / No", "yes_no", "is_residential"),
        ("Assigned Floors", False, "Comma-separated floor numbers, e.g. '1,2,3'", None, "assigned_floors"),
        ("Shift", False, "Pick from dropdown", "warden_shift", "shift"),
        ("Active", False, "Is this assignment currently active? Yes / No", "yes_no", "is_active"),
        ("Start Date", True, "Date assignment starts (YYYY-MM-DD)", None, "assigned_at"),
     ],
     [
        ["JKKN College of Engineering and Technology", "Dr. Ramesh Kumar", "EMP-2019-045", "BLK-A",
         "Warden", "9876543210", "Yes", "1,2", "Full-Time (Day + Night)", "Yes", "2026-04-20"],
     ]),

    ("5. Fee Config",
     "Fee structure per (Academic Year × Room Type × AC status). One row per combination.",
     [
        ("College Name", True, "Which college", INST, "institution_id"),
        ("Academic Year", True, "e.g. '2026-27'. Loader looks up the year UUID.", None, "_academic_year_label"),
        ("Room Type", True, "Pick from dropdown", "room_type", "room_type"),
        ("AC Status", True, "Pick from dropdown", "ac_status", "ac_status"),
        ("Annual Fee (₹)", True, "Total yearly fee", None, "annual_fee"),
        ("Semester Fee (₹)", False, "Half-year fee (usually annual ÷ 2)", None, "semester_fee"),
        ("Monthly Fee (₹)", False, "Monthly fee (usually annual ÷ 10)", None, "monthly_fee"),
        ("Refundable Deposit (₹)", True, "One-time refundable deposit", None, "deposit_amount"),
        ("Mess Fee Monthly (₹)", False, "Monthly mess charges", None, "mess_fee_monthly"),
        ("Mess Fee Semester (₹)", False, "Semester mess charges", None, "mess_fee_semester"),
        ("Electricity Charging", False, "Pick from dropdown", "electricity_charges", "electricity_charges"),
        ("Electricity Fixed Amount (₹)", False, "If 'Fixed monthly charge' above, amount", None, "electricity_fixed_amount"),
        ("Active", False, "Yes / No", "yes_no", "is_active"),
     ],
     [
        ["JKKN College of Engineering and Technology", "2026-27", "Double (2 beds)", "Non-AC",
         45000, 22500, 4500, 10000, 3500, 21000, "Included in hostel fees", "", "Yes"],
        ["JKKN College of Engineering and Technology", "2026-27", "Double (2 beds)", "AC",
         65000, 32500, 6500, 15000, 3500, 21000, "Fixed monthly charge", 500, "Yes"],
     ]),

    ("6. Leave Types",
     "Which leave types are allowed at each college + approval rules.",
     [
        ("College Name", True, "Which college", INST, "institution_id"),
        ("Leave Type", True, "Pick from dropdown", "leave_type", "leave_type"),
        ("Max Duration (days)", False, "Max days allowed in one request", None, "max_duration_days"),
        ("Parent Consent Required", False, "Yes / No", "yes_no", "requires_parent_consent"),
        ("Advance Notice (hours)", False, "Hours of advance notice required before leave starts", None, "advance_notice_hours"),
        ("Chief Warden Approval Required", False, "Yes / No", "yes_no", "requires_chief_warden"),
        ("Attachment Required", False, "e.g. medical certificate. Yes / No", "yes_no", "requires_attachment"),
        ("Active", False, "Is this leave type enabled? Yes / No", "yes_no", "is_active"),
     ],
     [
        ["JKKN College of Engineering and Technology", "Home Visit", 30, "Yes", 24, "No", "No", "Yes"],
        ["JKKN College of Engineering and Technology", "Medical", 15, "Yes", 0, "No", "Yes", "Yes"],
        ["JKKN College of Engineering and Technology", "Emergency", 7, "Yes", 0, "Yes", "No", "Yes"],
     ]),

    ("7. Mess Caterers",
     "Food vendors / in-house caterers contracted for the mess.",
     [
        ("College Name", True, "Which college", INST, "institution_id"),
        ("Caterer Name", True, "Business name", None, "name"),
        ("Owner Name", True, "Contact person", None, "owner_name"),
        ("Phone", True, "10-digit", None, "phone"),
        ("Email", False, "Business email", None, "email"),
        ("FSSAI License Number", False, "Food safety license", None, "fssai_license_number"),
        ("FSSAI Expiry Date", False, "YYYY-MM-DD", None, "fssai_expiry_date"),
        ("GST Number", False, "GSTIN", None, "gst_number"),
        ("Contract Start Date", True, "YYYY-MM-DD", None, "contract_start_date"),
        ("Contract End Date", True, "YYYY-MM-DD", None, "contract_end_date"),
        ("Contract Amount Monthly (₹)", False, "Monthly contract value", None, "contract_amount_monthly"),
        ("Billing Model", True, "Pick from dropdown", "billing_model", "billing_model"),
        ("Status", False, "Pick from dropdown (default: Active)", "caterer_status", "status"),
     ],
     [
        ["JKKN College of Engineering and Technology", "Amman Catering Services", "Ramesh Kumar", "9876501234",
         "amman@example.com", "12345678901234", "2027-03-31", "33ABCDE1234F1Z5",
         "2026-06-01", "2027-05-31", 250000, "Fixed Monthly Fee", "Active"],
     ]),

    ("8. Mess Menus",
     "Weekly menu per caterer. One row per (week × day × meal).",
     [
        ("College Name", True, "Which college", INST, "institution_id"),
        ("Caterer Name", True, "Must match a caterer from Mess Caterers sheet", None, "_caterer_name"),
        ("Block Code", False, "Specific block? Leave blank for all blocks", None, "_block_code"),
        ("Week Starting (Monday)", True, "YYYY-MM-DD — the Monday of that week", None, "week_start_date"),
        ("Day of Week", True, "Pick from dropdown", "day_of_week", "day_of_week"),
        ("Meal", True, "Pick from dropdown", "meal_type", "meal_type"),
        ("Menu Items", True, "Comma-separated e.g. 'Idli, Sambar, Chutney, Tea'", None, "items"),
        ("Special Items", False, "Comma-separated extras", None, "special_items"),
        ("Dietary Tags", False, "Comma-separated e.g. 'vegetarian, jain'", None, "dietary_tags"),
        ("Cost per Plate (₹)", False, "Estimated cost", None, "estimated_cost_per_plate"),
        ("Special Day", False, "Festival / special occasion? Yes / No", "yes_no", "is_special_day"),
        ("Special Day Name", False, "e.g. 'Pongal', 'Diwali'", None, "special_day_name"),
        ("Status", False, "Pick from dropdown (default: Planned)", "menu_status", "status"),
     ],
     [
        ["JKKN College of Engineering and Technology", "Amman Catering Services", "", "2026-06-01",
         "Monday", "Breakfast", "Idli, Sambar, Chutney, Tea", "", "vegetarian", 35, "No", "", "Planned"],
        ["JKKN College of Engineering and Technology", "Amman Catering Services", "", "2026-06-01",
         "Monday", "Lunch", "Rice, Sambar, Rasam, Poriyal, Curd", "", "vegetarian", 60, "No", "", "Planned"],
     ]),

    ("9. Maintenance SLA",
     "Response time targets (how fast maintenance issues must be resolved).",
     [
        ("College Name", True, "Which college", INST, "institution_id"),
        ("Maintenance Category", True, "Pick from dropdown", "maintenance_category", "category"),
        ("Priority", True, "Pick from dropdown", "maintenance_priority", "priority"),
        ("Target Resolution (hours)", True, "How many hours to fix", None, "sla_hours"),
        ("Escalate After (hours)", False, "If unresolved after this many hours, escalate", None, "escalation_after_hours"),
        ("Escalate To (role)", False, "e.g. 'Chief Warden', 'Administrator'", None, "escalation_to_role"),
        ("Active", False, "Yes / No", "yes_no", "is_active"),
     ],
     [
        ["JKKN College of Engineering and Technology", "Electrical", "Critical (fix immediately)", 4, 2, "Chief Warden", "Yes"],
        ["JKKN College of Engineering and Technology", "Plumbing", "High (same day)", 24, 12, "Warden", "Yes"],
        ["JKKN College of Engineering and Technology", "Cleaning", "Low (within a week)", 72, 48, "Warden", "Yes"],
     ]),

    ("10. Cleaning Schedules",
     "Recurring housekeeping schedules per block / floor.",
     [
        ("College Name", True, "Which college", INST, "institution_id"),
        ("Block Code", False, "Leave blank for all blocks in the college", None, "_block_code"),
        ("Floor Number", False, "Leave blank for all floors in the block", None, "floor_number"),
        ("Cleaning Type", True, "Pick from dropdown", "cleaning_type", "cleaning_type"),
        ("Frequency", True, "Pick from dropdown", "cleaning_frequency", "frequency"),
        ("Scheduled Time", False, "HH:MM (24-hour), e.g. '06:00'", None, "scheduled_time"),
        ("Assigned Staff Name", False, "Cleaning staff name", None, "assigned_staff"),
        ("Assigned Staff Phone", False, "10-digit mobile", None, "assigned_staff_phone"),
        ("Active", False, "Yes / No", "yes_no", "is_active"),
     ],
     [
        ["JKKN College of Engineering and Technology", "BLK-A", 1, "Daily Sweeping", "Daily", "06:00", "Selvi R", "9876543299", "Yes"],
        ["JKKN College of Engineering and Technology", "BLK-A", "", "Water Tank Cleaning", "Monthly", "08:00", "Kumar S", "9876543298", "Yes"],
     ]),

    ("11. Laundry Config",
     "Laundry service setup per block (or institution-wide if Block Code is blank).",
     [
        ("College Name", True, "Which college", INST, "institution_id"),
        ("Block Code", False, "Leave blank for institution-wide", None, "_block_code"),
        ("Service Type", False, "Pick from dropdown", "laundry_service_type", "service_type"),
        ("Vendor Name", False, "External vendor name", None, "vendor_name"),
        ("Vendor Phone", False, "10-digit", None, "vendor_phone"),
        ("Contract Start", False, "YYYY-MM-DD", None, "vendor_contract_start"),
        ("Contract End", False, "YYYY-MM-DD", None, "vendor_contract_end"),
        ("Collection Days", False, "Comma-separated, e.g. 'Monday, Thursday'", None, "collection_days"),
        ("Delivery Days", False, "Comma-separated, e.g. 'Wednesday, Saturday'", None, "delivery_days"),
        ("Max Items per Student", False, "Integer", None, "max_items_per_student"),
        ("Cost per Item (₹)", False, "Charge per clothing item", None, "cost_per_item"),
        ("Included in Hostel Fees", False, "Yes / No", "yes_no", "is_included_in_fees"),
        ("Active", False, "Yes / No", "yes_no", "is_active"),
     ],
     [
        ["JKKN College of Engineering and Technology", "BLK-A", "External Vendor", "Clean&Fresh Laundry", "9876512345",
         "2026-06-01", "2027-05-31", "Monday, Thursday", "Wednesday, Saturday", 15, 8, "No", "Yes"],
     ]),

    ("12. Safety Equipment",
     "Fire extinguishers, CCTV, smoke detectors, first-aid kits, etc.",
     [
        ("College Name", True, "Which college", INST, "institution_id"),
        ("Block Code", False, "Leave blank if institution-wide equipment", None, "_block_code"),
        ("Equipment Type", True, "e.g. 'Fire Extinguisher', 'CCTV', 'Smoke Detector', 'First Aid Kit'", None, "equipment_type"),
        ("Identifier/Name", True, "Unique label e.g. 'FE-A1-01'", None, "name"),
        ("Location", False, "Where it is e.g. 'Floor 1 stairwell'", None, "location"),
        ("Serial Number", False, "Serial / model number", None, "serial_number"),
        ("Last Inspection Date", False, "YYYY-MM-DD", None, "last_inspection_date"),
        ("Next Inspection Date", False, "YYYY-MM-DD", None, "next_inspection_date"),
        ("Status", False, "e.g. 'Operational', 'Needs Service', 'Out of Service'", None, "status"),
        ("Condition", False, "e.g. 'Good', 'Fair', 'Poor'", None, "condition"),
        ("Installed Date", False, "YYYY-MM-DD", None, "installed_date"),
        ("Warranty Expiry", False, "YYYY-MM-DD", None, "warranty_expiry"),
     ],
     [
        ["JKKN College of Engineering and Technology", "BLK-A", "Fire Extinguisher", "FE-A1-01",
         "Floor 1 stairwell", "FE-2024-001", "2026-01-15", "2026-07-15", "Operational", "Good",
         "2024-06-01", "2027-06-01"],
     ]),
]

# Sheets that need IT assistance (flagged in README, kept out of the main fill-flow)
ADVANCED_NOTE = (
    "The following setup items need technical assistance and should NOT be filled by wardens:\n"
    "• Allocations (assigning learners to beds) — use the /campus-living/allocations/new UI, one learner at a time\n"
    "• Alert Rules — configured by IT team via the Alerts settings page\n"
    "• Onboarding Checklists — configured via /campus-living/allocations/onboarding/templates\n"
    "• Community config, Pulse surveys, Emergency contacts — configured via admin settings"
)


# --- Helpers ---------------------------------------------------------------

def register_named_ranges(wb):
    """Kept for Excel convenience; Google Sheets ignores it so we also use inline refs."""
    last_row = 1 + len(INSTITUTIONS)
    ref = f"'REF Institutions'!$A$2:$A${last_row}"
    wb.defined_names["institutions"] = DefinedName("institutions", attr_text=ref)


def apply_validations(ws, columns, data_start_row=4, data_end_row=1003):
    """Write data validations that work in BOTH Excel and Google Sheets.

    Rules for cross-compat:
    - Inline lists (enums, Yes/No): formula1='"A,B,C"' — works everywhere
    - Range refs (college list): formula1="'REF Institutions'!$A$2:$A$9" WITHOUT leading '='
      (Excel tolerates either; Google Sheets drops the validation if it sees '=' prefix
       or an unresolved named range)
    - Named ranges (=institutions): DROPPED — Google Sheets import strips these
    """
    last_inst_row = 1 + len(INSTITUTIONS)
    inst_range = f"'REF Institutions'!$A$2:$A${last_inst_row}"

    for col_idx, (label, _req, _hint, validation_key, _db) in enumerate(columns, 1):
        if not validation_key:
            continue
        letter = get_column_letter(col_idx)
        cell_range = f"{letter}{data_start_row}:{letter}{data_end_row}"

        if validation_key == INST:
            # Direct cross-sheet range ref — both Excel and Google Sheets accept
            dv = DataValidation(
                type="list",
                formula1=inst_range,
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid college",
                error="Pick a JKKN college from the dropdown.",
            )
        elif validation_key in LABELS:
            values = [label for label, _db in LABELS[validation_key]]
            formula = '"' + ",".join(values) + '"'
            if len(formula) > 255:
                continue
            dv = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid choice",
                error=f"Pick one of: {', '.join(values)}",
            )
        else:
            continue

        dv.add(cell_range)
        ws.add_data_validation(dv)


def build_readme(wb):
    ws = wb.create_sheet("README", 0)
    ws["A1"] = "MyJKKN Campus Living — Bulk Setup Template"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28

    lines = [
        ("", None),
        ("Who Fills This", TITLE_FONT),
        ("Hostel wardens and admin staff. No technical knowledge needed — every field either has a dropdown or a plain-English hint.", None),
        ("", None),
        ("What It Does", TITLE_FONT),
        ("One-time setup of the entire Campus Living module for a college. Fill the sheets in order, hand the file back to IT, they run one command that loads everything into the system.", None),
        ("", None),
        ("The 3-Step Flow", TITLE_FONT),
        ("  Step 1 — Fill sheets 1, 2, 3 first (Blocks → Rooms → Beds). This is the physical hierarchy.", None),
        ("  Step 2 — Fill the config sheets (4–12). Order within this group doesn't matter.", None),
        ("  Step 3 — Save the file and send it to IT. They run: python3 scripts/load-campus-living-template.py <your-file>", None),
        ("", None),
        ("Color Legend", TITLE_FONT),
        ("  🟥 RED header = required (the row will be rejected if this is blank)", None),
        ("  ⬜ GRAY header = optional", None),
        ("  ⬛ DARK GRAY row 3 = hint / example / format", None),
        ("  🟨 YELLOW rows = example data — REPLACE or DELETE these before saving", None),
        ("  🟩 GREEN rows = empty — TYPE YOUR DATA HERE", None),
        ("", None),
        ("How Dropdowns Work", TITLE_FONT),
        ("  • Click any green cell in a dropdown column — a ▼ arrow appears on the right side", None),
        ("  • Click ▼ and pick from the list. You can also start typing and it will autocomplete.", None),
        ("  • Dropdowns use plain English ('Boys Hostel', 'AC', 'Yes') — not database codes.", None),
        ("  • Typing something outside the list → Excel rejects the entry with an error.", None),
        ("", None),
        ("How Sheets Link Together", TITLE_FONT),
        ("  You will NEVER see a UUID. You link sheets using human-readable codes:", None),
        ("    • Blocks sheet defines a 'Block Code' (e.g. 'BLK-A')", None),
        ("    • Rooms sheet references that 'Block Code' + gives each room a 'Room Number'", None),
        ("    • Beds sheet references 'Block Code' + 'Room Number' + gives each bed a 'Bed Number'", None),
        ("    • Wardens sheet references 'Block Code' + 'Employee ID' (from HR system)", None),
        ("    • Fee Config references 'Academic Year' (e.g. '2026-27')", None),
        ("  The IT loader script converts all these to internal UUIDs at import time.", None),
        ("", None),
        ("What's NOT in This Template (need IT assistance)", TITLE_FONT),
        (ADVANCED_NOTE, None),
        ("", None),
        ("Daily-use data (visitors, attendance, meal bookings, gate passes, maintenance requests, leave requests, etc.)", None),
        ("is NOT seeded through this template. It's generated naturally as learners use the app.", None),
        ("", None),
        ("Format Rules", TITLE_FONT),
        ("  • Dates: YYYY-MM-DD (e.g. 2026-06-01). No DD/MM/YYYY, no DD-Mon-YY.", None),
        ("  • Times: HH:MM 24-hour (e.g. 22:00 = 10 PM, 06:00 = 6 AM).", None),
        ("  • Phone: 10 digits, no spaces, no country code (e.g. 9876543210).", None),
        ("  • Money: Numbers only, no ₹ symbol, no commas (e.g. 45000 not '45,000' or '₹45,000').", None),
        ("  • Comma-separated lists: separate with commas + space (e.g. 'Idli, Sambar, Chutney').", None),
        ("", None),
        ("Need Help?", TITLE_FONT),
        ("  • If a field is confusing, leave it blank (if it's optional) or ask IT.", None),
        ("  • Fill one row completely first, send to IT to test, then fill the rest.", None),
        ("  • Don't delete the header row (row 2) or hint row (row 3). Only fill the GREEN rows.", None),
    ]
    for i, (text, font) in enumerate(lines, 3):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
    ws.column_dimensions["A"].width = 120


def build_ref_institutions(wb):
    """Lists college names only — UUIDs hidden in a ref column but not shown in dropdowns."""
    ws = wb.create_sheet("REF Institutions")
    headers = ["College Name", "Type"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = BLUE
        c.font = WHITE_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    type_lookup = {
        "JKKN College of Engineering and Technology": "Autonomous",
        "JKKN College of Pharmacy": "Autonomous",
        "JKKN Dental College and Hospital": "Autonomous",
        "JKKN College of Allied Health Sciences": "Autonomous",
        "JKKN College of Nursing and Research": "Autonomous",
        "JKKN College of Arts and Science (Aided)": "Aided (Govt)",
        "JKKN College of Arts and Science (Self)": "Self-Financed",
        "JKKN College of Education": "Self-Financed",
    }
    for r, name in enumerate(INSTITUTIONS, 2):
        ws.cell(row=r, column=1, value=name).border = BORDER
        ws.cell(row=r, column=2, value=type_lookup.get(name, "")).border = BORDER
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 18
    ws.freeze_panes = "A2"


def build_ref_dropdowns(wb):
    """Shows the human labels for each dropdown — reference only."""
    ws = wb.create_sheet("REF Dropdowns")
    ws.cell(row=1, column=1, value="Dropdown Field").fill = BLUE
    ws.cell(row=1, column=2, value="Allowed Values").fill = BLUE
    for c in (1, 2):
        cell = ws.cell(row=1, column=c)
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    r = 2
    for key, pairs in LABELS.items():
        ws.cell(row=r, column=1, value=key.replace("_", " ").title()).border = BORDER
        ws.cell(row=r, column=2, value=" | ".join(lbl for lbl, _ in pairs)).border = BORDER
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100
    ws.freeze_panes = "A2"


def build_data_sheet(wb, sheet_name, description, columns, sample_rows):
    ws = wb.create_sheet(sheet_name)

    # Row 1: description
    ws.cell(row=1, column=1, value=f"ℹ  {description}").font = Font(color="1E3A8A", italic=True, size=10)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.row_dimensions[1].height = 22

    # Row 2: headers
    for i, (label, required, _hint, _validation, _db) in enumerate(columns, 1):
        c = ws.cell(row=2, column=i, value=("★ " if required else "") + label)
        c.fill = RED if required else GRAY_HEAD
        c.font = WHITE_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[2].height = 40

    # Row 3: hints
    for i, (_label, _required, hint, _validation, _db) in enumerate(columns, 1):
        c = ws.cell(row=3, column=i, value=hint)
        c.fill = GRAY_LIGHT
        c.font = HINT_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[3].height = 48

    # Sample rows in yellow (rows 4..4+N-1)
    sample_count = len(sample_rows)
    for ri, row in enumerate(sample_rows, 4):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = YELLOW
            c.font = Font(size=10)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c.border = BORDER

    # Green input rows below
    input_start = 4 + sample_count
    input_end = 1003
    for r in range(input_start, input_end + 1):
        for ci in range(1, len(columns) + 1):
            c = ws.cell(row=r, column=ci)
            c.fill = GREEN
            c.border = BORDER
            c.font = Font(size=10)

    # Column widths
    for i, (label, _req, hint, _v, _d) in enumerate(columns, 1):
        width = max(len(label) + 6, min(len(hint) // 2, 28), 16)
        width = min(width, 42)
        ws.column_dimensions[get_column_letter(i)].width = width

    # Dropdowns
    apply_validations(ws, columns)

    ws.freeze_panes = "A4"


def main():
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    build_readme(wb)
    build_ref_institutions(wb)
    build_ref_dropdowns(wb)
    register_named_ranges(wb)

    for sheet_name, desc, cols, samples in SHEETS:
        build_data_sheet(wb, sheet_name, desc, cols, samples)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)

    total_dropdowns = sum(1 for _, _, cols, _ in SHEETS for c in cols if c[3])
    print(f"✅ Wrote {OUT}")
    print(f"   Sheets: README + REF Institutions + REF Dropdowns + {len(SHEETS)} data-entry = {3 + len(SHEETS)} total")
    print(f"   Dropdowns (all plain English): {total_dropdowns}")
    print(f"   Zero UUIDs shown to user. ✅")


if __name__ == "__main__":
    main()
