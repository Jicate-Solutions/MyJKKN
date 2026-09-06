#!/usr/bin/env python3
"""Convert a parsed AICTE/Anna University curriculum into a BoS courses import .xlsx.

The PDF -> structured-JSON step is done by Claude (it varies per document).
This script does the deterministic part: apply the fixed mapping rules, write the
template preserving its Lists/Instructions sheets (and dropdown validation), and
report a reconciliation summary so a dropped/duplicated course is caught.

Usage:
    python convert_curriculum.py <courses.json>

courses.json schema:
{
  "template": "path/to/bos-courses-import-template.xlsx",   # required
  "output":   "path/to/output.xlsx",                        # required
  "expected_total_credits": 162,                            # optional, for reconciliation
  "defaults": {"internal": 25, "external": 75},             # optional, per-course can override
  "courses": [
    {"code":"HS3152","name":"Professional English - I","aicte":"HSMC",
     "L":3,"T":0,"P":0,"credits":3}
    // optional per-course keys: "category" (override), "exam", "internal", "external"
  ]
}

Rules applied (see references/mapping.md for the rationale):
- aicte code  -> Type column           via AICTE_TO_TYPE
- L/T/P split -> Category column        via derive_category()   (unless "category" given)
- exam:   null -> 0 if credits==0 else 3
- internal/external default 25/75 (override via defaults or per-course)
"""
import json
import sys

# AICTE category code -> template "Type" value. MUST be a value present on the
# Lists sheet, else the import dialog rejects the row. Verified at runtime too.
AICTE_TO_TYPE = {
    "HSMC": "Humanities, Social Sciences & Management Courses",
    "BSC":  "Basic Science Courses",
    "ESC":  "Engineering Science Courses",
    "PCC":  "Professional Core Courses",
    "PEC":  "Programme Elective",            # Lists has no literal "Professional Elective"
    "OEC":  "Open Elective Courses",
    "EEC":  "Employability Enhancement Courses",
    "MC":   "Mandatory Courses",
    "":     "",                              # uncategorised (e.g. Induction Programme)
}


def derive_category(L, T, P):
    """L-T-P weekly period split -> template 'Category' (nature of the course)."""
    L, T, P = int(L or 0), int(T or 0), int(P or 0)
    if P > 0 and (L > 0 or T > 0):
        return "Theory + Practical"
    if P > 0 and L == 0 and T == 0:
        return "Practical"
    return "Theory"


def load_valid_lists(wb):
    ws = wb["Lists"]
    cats, types = set(), set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            cats.add(row[0])
        if row and len(row) > 1 and row[1]:
            types.add(row[1])
    return cats, types


def main():
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(2)

    import openpyxl  # imported here so --help works without the dep

    with open(sys.argv[1], encoding="utf-8") as f:
        spec = json.load(f)

    template = spec["template"]
    output = spec["output"]
    defaults = spec.get("defaults", {})
    def_internal = defaults.get("internal", 25)
    def_external = defaults.get("external", 75)

    wb = openpyxl.load_workbook(template)
    ws = wb["Courses"]
    valid_cats, valid_types = load_valid_lists(wb)

    # clear existing data rows, keep header (row 1)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    problems, seen_codes = [], set()
    total_credits = 0.0
    r = 2
    for c in spec["courses"]:
        code = str(c["code"]).strip()
        name = str(c["name"]).strip()
        aicte = (c.get("aicte") or "").strip().upper()
        L, T, P = c.get("L", 0), c.get("T", 0), c.get("P", 0)
        credits = c.get("credits", 0) or 0

        category = c.get("category") or derive_category(L, T, P)
        typ = AICTE_TO_TYPE.get(aicte)
        if typ is None:
            problems.append(f"{code}: unknown AICTE code '{aicte}' (no Type mapping)")
            typ = ""
        exam = c.get("exam")
        if exam is None:
            exam = 0 if credits == 0 else 3
        internal = c.get("internal", def_internal)
        external = c.get("external", def_external)
        theory_hours = int(L or 0) + int(T or 0)
        practical_hours = int(P or 0)

        # validations
        if not code.isalnum():
            problems.append(f"{code}: course code must be letters/digits only")
        if code in seen_codes:
            problems.append(f"{code}: duplicate course code")
        seen_codes.add(code)
        if category not in valid_cats:
            problems.append(f"{code}: Category '{category}' not in Lists sheet")
        if typ and typ not in valid_types:
            problems.append(f"{code}: Type '{typ}' not in Lists sheet")

        ws.cell(r, 1, code)
        ws.cell(r, 2, name)
        ws.cell(r, 3, category)
        ws.cell(r, 4, typ)
        ws.cell(r, 5, exam)
        ws.cell(r, 6, credits)
        ws.cell(r, 7, theory_hours)
        ws.cell(r, 8, practical_hours)
        ws.cell(r, 9, internal)
        ws.cell(r, 10, external)
        total_credits += float(credits)
        r += 1

    wb.save(output)

    n = len(spec["courses"])
    print(f"Wrote {output}")
    print(f"  Courses: {n} (rows 2..{r - 1})")
    print(f"  Sum of credits: {total_credits:g}")
    exp = spec.get("expected_total_credits")
    if exp is not None:
        diff = total_credits - exp
        if abs(diff) < 1e-9:
            print(f"  Reconciles to expected_total_credits ({exp}). OK")
        else:
            print(f"  NOTE: differs from expected_total_credits ({exp}) by {diff:g} "
                  f"- confirm the residual is explainable (e.g. NCC/extra-credit rows).")
    if problems:
        print("\nVALIDATION ISSUES (fix before importing):")
        for p in problems:
            print("  -", p)
        sys.exit(1)
    print("\nAll Category/Type values valid against the Lists sheet.")


if __name__ == "__main__":
    main()
