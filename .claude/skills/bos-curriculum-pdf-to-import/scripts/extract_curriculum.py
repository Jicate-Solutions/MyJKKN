#!/usr/bin/env python3
"""Extract courses from a curriculum file (any known JKKN format) and write the
BoS import .xlsx directly — auto-detecting the layout. Use this when the source is a
machine-readable table (.docx/.pdf/.rtf); it removes the manual JSON step that
convert_curriculum.py needs.

Usage:
    python extract_curriculum.py <curriculum-file> [--template T.xlsx] [--out O.xlsx]
                                 [--expected-credits N]

Handles four formats (see references/mapping.md → "Source formats"):
  A. R-2021 columnar   .docx/.pdf  — CODE|TITLE|CATEGORY|L|T|P|TCP|CREDITS
  B. R-2026 REC-26     .docx/.pdf  — CODE|Name|CourseType|L-T-P(combined)|TCP|Credits|Category
  C. R-2025 hybrid     .docx/.pdf  — CourseType + separate OR combined L-T-P + trailing Category
  D. R-2021 RTF        .rtf        — striprtf pipe-delimited rows (same fields as A)
  (Anna University NON-AUTONOMOUS affiliated text-layout PDFs are NOT auto-handled —
   they have no ruled tables and omit codes from Sem III on; parse those manually.)

Deps: openpyxl (required), python-docx (.docx), pdfplumber (.pdf), striprtf (.rtf).
"""
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from convert_curriculum import AICTE_TO_TYPE, derive_category, load_valid_lists

# --- new-regulation (R-2025 / R-2026) category code -> template Type (Lists value) ---
NEWCAT_TYPE = {
    "HS": "Humanities, Social Sciences & Management Courses",
    "HUM": "Humanities, Social Sciences & Management Courses",
    "HSS": "Humanities, Social Sciences & Management Courses",
    "BS": "Basic Science", "BSS": "Basic Science",
    "ES": "Engineering Science (General)", "GES": "Engineering Science (General)",
    "SD": "Skill Development", "EEC": "Employability Enhancement Courses",
    "PC": "Professional Core Courses", "PE": "Programme Elective",
    "OE": "Open Elective Courses", "MC": "Mandatory Courses",
    "PCS": "Professional Competency Skill", "NEC": "Non Major Elective",
    "PW": "Project Work", "PROJ": "Project Work", "SL": "Self Learning",
    "SA": "Mandatory Courses",  # co-curricular NCC/NSS/NSO, non-credit
}
# CourseType (new formats) -> template Category, but L-T-P is authoritative when present.
CTYPE_CAT = {"P": "Practical", "L": "Practical"}

# merged code->Type lookup (AICTE keys and new-reg keys don't collide)
TYPE_LOOKUP = {**AICTE_TO_TYPE, **NEWCAT_TYPE}
SECTION = {"THEORY", "PRACTICALS", "PRACTICAL", "TOTAL", "TOTAL CREDITS", "MANDATORY COURSES",
           "AUDIT COURSE", "HONOURS", "MINOR", "ELECTIVE", "ONE CREDIT COURSES",
           "VALUE ADDED COURSES", "PROFESSIONAL ELECTIVE", "OPEN ELECTIVE"}
# real course code: alphanumeric, has letters AND digits, 3-10 chars. Covers IT3401,
# MA25C01, 26HSS01 (year-prefixed), NCC1 placeholders — rejects T/LIT/SEMESTERII/section labels.
CODE_RE = re.compile(r"^(?=.*[A-Z])(?=.*\d)[A-Z0-9]{3,10}$")
# stricter — RTF parses raw pipe-text (no header-based table isolation), so prose rows like
# "PO2"/"PSO1"/"C01" would slip past CODE_RE. Real Anna-Univ codes are 2-4 letters + 3-4 digits.
CODE_RE_RTF = re.compile(r"^[A-Z]{2,4}\d{3,4}[A-Z]?\d?$")


def _clean(s):
    return re.sub(r"\s+", " ", (s or "").strip())


def _num(s):
    s = _clean(s)
    if not s or s == "-":
        return None
    if re.search(r"non\s*-?\s*credit", s, re.I):
        return 0
    m = re.search(r"\d+(\.\d+)?", s)
    return float(m.group()) if m else None


# ------------------------------------------------------------------ table readers
def _tables_docx(p):
    import docx
    return [[[_clean(c.text) for c in r.cells] for r in t.rows] for t in docx.Document(p).tables]


def _tables_pdf(p):
    import pdfplumber
    out = []
    with pdfplumber.open(p) as pdf:
        for page in pdf.pages:
            for tb in page.extract_tables():
                out.append([[_clean(c) for c in row] for row in tb])
    return out


def _read_tables(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".docx":
        return _tables_docx(path)
    if ext == ".pdf":
        return _tables_pdf(path)
    return []


# ------------------------------------------------------------------ format detection
def _is_new_format(tables):
    """New/hybrid layouts carry a 'Course Type' column plus a 'Category' column."""
    for rows in tables:
        for r in rows[:4]:
            # strip ALL whitespace: pdfplumber often splits header words mid-token
            # ("Cours\ne\nTyp\ne" -> "COURS E TYP E"), which would hide TYPE/CATEGORY.
            j = re.sub(r"\s+", "", " ".join(r).upper())
            if ("CODE" in j) and ("TYPE" in j) and ("CATEGORY" in j):
                return True
    return False


# ------------------------------------------------------------------ parser A/D: columnar (R-2021)
def _parse_columnar(tables):
    """CODE|TITLE|CATEGORY|L|T|P|TOTAL CONTACT|CREDITS. Returns unified dicts."""
    courses, ncc = [], 0
    for rows in tables:
        cols = _find_columnar_cols(rows)
        if not cols:
            continue
        code_i, cate_i, L_i, T_i, P_i, cr_i = cols
        for row in rows:
            if not any(row):
                continue
            rj = " ".join(row).upper()
            if "COURSE CODE" in rj or "COURSE TITLE" in rj or "CATE-GORY" in rj or "CATE GORY" in rj:
                continue
            f0 = _clean(row[0]).upper().rstrip(".")
            ne = [c for c in row if _clean(c)]
            if ne and len(set(ne)) == 1 and ne[0].upper() in SECTION:
                continue
            if f0 in SECTION:
                continue

            def g(i):
                return _clean(row[i]) if i < len(row) else ""
            code = g(code_i)
            cate = g(cate_i)
            if cate.upper() in SECTION:
                continue
            has_serial = bool(re.match(r"^\d+\.?$", _clean(row[0])))
            if not has_serial and not code:
                continue
            name_parts = [g(j) for j in range(code_i + 1, cate_i) if g(j)]
            name = re.sub(r"\s*[\$#\*]+\s*$", "", name_parts[0]).strip() if name_parts else ""
            L = int(_num(g(L_i)) or 0)
            T = int(_num(g(T_i)) or 0)
            P = int(_num(g(P_i)) or 0)
            cr = _num(g(cr_i))
            if not code and "NCC" in name.upper():
                ncc += 1
                code = f"NCC{ncc}"
            code = re.sub(r"[^A-Za-z0-9]", "", code)
            if not code or not name:
                continue
            courses.append({"code": code, "name": name,
                            "type_code": cate.upper().replace("-", "").replace(" ", ""),
                            "L": L, "T": T, "P": P, "credits": cr if cr is not None else 0,
                            "ctype": None})
    return courses


def _find_columnar_cols(rows):
    for row in rows[:3]:
        up = [c.upper() for c in row]
        j = " ".join(up)
        if "CODE" not in j:
            continue
        tot = None
        for i, c in enumerate(up):
            if "TOTAL" in c and "CONTACT" in c:
                tot = i
                break
        if tot is None:
            for i, c in enumerate(up):
                if "CREDIT" in c:
                    tot = i - 1
                    break
        if tot is None or tot < 4:
            continue
        code = next((i for i, c in enumerate(up) if "CODE" in c), 1)
        return (code, tot - 4, tot - 3, tot - 2, tot - 1, tot + 1 if tot + 1 < len(row) else len(row) - 1)
    return None


# ------------------------------------------------------------------ parser B/C: new + hybrid
def _norm_cat(s):
    s = _clean(s).upper()
    base = re.split(r"[\(\-–/]", s)[0].strip().replace(" ", "")
    m = re.match(r"[A-Z]+", base)
    base = m.group() if m else ""
    # R-2025 writes Engineering Science courses as "ES (PC)" / "ES (PE)" / "ES (G)";
    # honor the qualifier so core reads as core and electives read as electives.
    # (Only ES is qualified this way — every other code keeps its bare mapping.)
    if base == "ES":
        q = re.search(r"\(\s*([A-Z]{1,3})\s*\)", s)
        if q and q.group(1) in ("PC", "PE", "OE"):
            return q.group(1)
    return base


def _parse_new(tables):
    courses = []
    for rows in tables:
        idx = _find_new_cols(rows)
        if not idx:
            continue
        for row in rows:
            if not any(row):
                continue
            rj = " ".join(row).upper()
            if "COURSE CODE" in rj or "COURSE NAME" in rj or "COURSE TITLE" in rj or "L-T-P" in rj:
                continue
            f0 = _clean(row[0]).upper().rstrip(".")
            ne = [_clean(c) for c in row if _clean(c)]
            if ne and len(set(ne)) == 1 and (ne[0].upper() in SECTION or "SEMESTER" in ne[0].upper()):
                continue
            if f0 in SECTION or "SEMESTER" in f0:
                continue

            def g(i):
                return _clean(row[i]) if (i is not None and i < len(row)) else ""
            code = re.sub(r"[^A-Za-z0-9]", "", g(idx["code"]))
            name = g(idx["name"])
            if not code or not name or name.upper() in ("COURSE NAME", "COURSE TITLE"):
                continue
            ltp = g(idx["ltp"])
            if not re.match(r"^\s*\d+\s*-\s*\d+\s*-\s*\d+", ltp):
                ltp = ""
                for c in row:
                    if re.match(r"^\s*\d+\s*-\s*\d+\s*-\s*\d+\s*$", _clean(c)):
                        ltp = _clean(c)
                        break
            m = re.match(r"^\s*(\d+)\s*-\s*(\d+)\s*-\s*(\d+)", ltp)
            if m:
                L, T, P = int(m.group(1)), int(m.group(2)), int(m.group(3))
            else:
                L = int(_num(g(idx["L"])) or 0)
                T = int(_num(g(idx["T"])) or 0)
                P = int(_num(g(idx["P"])) or 0)
            cr = _num(g(idx["credits"]))
            ctype = g(idx["ctype"]).upper().replace("-", "")
            base = _norm_cat(g(idx["cate"]))
            if not CODE_RE.match(code):
                continue  # drop section/junk rows without a real course code
            courses.append({"code": code,
                            "name": re.sub(r"\s*[\$#\*\^]+\s*$", "", name).strip(),
                            "type_code": base, "L": L, "T": T, "P": P,
                            "credits": cr if cr is not None else 0, "ctype": ctype})
    return courses


def _find_new_cols(rows):
    hdr = [r for r in rows[:4] if "CODE" in " ".join(r).upper()
           and ("NAME" in " ".join(r).upper() or "TITLE" in " ".join(r).upper())]
    if not hdr:
        return None
    ncol = max(len(r) for r in hdr)
    kw = [""] * ncol
    for r in rows[:4]:
        for i, c in enumerate(r):
            if i < ncol:
                kw[i] = (kw[i] + " " + c.upper()).strip()

    def find(pred, source=kw):
        for i, k in enumerate(source):
            if pred(k):
                return i
        return None
    # whitespace-free per-column keywords, so split header words ("Cours\ne\nTyp\ne"
    # -> "COURSETYPE", "Categor\ny" -> "CATEGORY") still match substrings. The single-
    # letter L/T/P finds keep the spaced tokens (they live on their own header sub-row).
    kws = [re.sub(r"\s+", "", k) for k in kw]
    tok = lambda k, t: t in k.split()
    idx = {
        "code": find(lambda k: "CODE" in k, kws),
        "name": find(lambda k: "NAME" in k or "TITLE" in k, kws),
        "ctype": find(lambda k: "TYPE" in k, kws),
        "credits": find(lambda k: "CREDIT" in k, kws),
        "cate": find(lambda k: "CATEGORY" in k or k == "CAT", kws),
        "ltp": find(lambda k: "L-T" in k or "LTP" in k, kws),
        "L": find(lambda k: tok(k, "L") and "CREDIT" not in k and "TCP" not in k),
        "T": find(lambda k: tok(k, "T")),
        "P": find(lambda k: tok(k, "P")),
    }
    if idx["code"] is None or idx["name"] is None or idx["cate"] is None:
        return None
    return idx


# ------------------------------------------------------------------ parser D: RTF
def _parse_rtf(path):
    from striprtf.striprtf import rtf_to_text
    txt = rtf_to_text(open(path, encoding="latin-1").read())
    ROW = re.compile(r"^\d+\.?\|")
    STOP = re.compile(r"^(SEMESTER|S\.No|For\b|\*|Total|VERTICAL|PROFESSIONAL ELECTIVE|"
                      r"OPEN ELECTIVE|MANDATORY|HONOURS|MINOR)", re.I)
    rows, buf = [], None
    for l in txt.splitlines():
        s = l.strip()
        if ROW.match(s):
            if buf:
                rows.append(buf)
            buf = s
        elif buf is not None:
            if not s or STOP.match(s) or ("|" not in s and len(s) > 60):
                rows.append(buf)
                buf = None
                continue
            buf += " " + s
    if buf:
        rows.append(buf)
    courses = []
    for r in rows:
        p = [x.strip() for x in r.split("|")]
        if len(p) < 9:
            continue
        code = re.sub(r"[^A-Za-z0-9]", "", p[1])
        if not CODE_RE_RTF.match(code):
            continue
        name = re.sub(r"\s*[\$#\*]+\s*$", "", p[2]).strip()
        if not name:
            continue
        courses.append({"code": code, "name": name,
                        "type_code": p[3].upper().replace("-", "").replace(" ", ""),
                        "L": int(_num(p[4]) or 0), "T": int(_num(p[5]) or 0),
                        "P": int(_num(p[6]) or 0),
                        "credits": _num(p[8]) if _num(p[8]) is not None else 0, "ctype": None})
    return courses


# ------------------------------------------------------------------ orchestration
def extract(path):
    """Return (courses, format_name). Dedups by code (keeps first)."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".rtf":
        raw, fmt = _parse_rtf(path), "D:rtf-columnar"
    else:
        tables = _read_tables(path)
        if not tables:
            return [], "none (no extractable tables — affiliated text-layout PDF? parse manually)"
        if _is_new_format(tables):
            raw, fmt = _parse_new(tables), "B/C:coursetype+category"
        else:
            raw, fmt = _parse_columnar(tables), "A:columnar"
    seen, out, ncc = set(), [], 0
    for c in raw:
        code = c["code"]
        if re.match(r"^NCC\d+$", code):  # renumber NCC globally so merged files don't collide
            ncc += 1
            code = c["code"] = f"NCC{ncc}"
        if code in seen:
            continue
        seen.add(code)
        out.append(c)
    return out, fmt


def to_rows(courses):
    """Map unified course dicts to the 10 template columns."""
    rows, problems = [], []
    for c in courses:
        nm = re.sub(r"^\s*/", "", c["name"]).strip()
        low = nm.lower()
        if any(k in low for k in ("project work", "internship", "mini project",
                                  "summer intern", "dissertation")):
            cat = "Project"
        elif c.get("ctype") in ("P", "L"):
            cat = "Practical"
        else:
            cat = derive_category(c["L"], c["T"], c["P"])
        typ = TYPE_LOOKUP.get(c["type_code"])
        if typ is None:
            if c["type_code"]:
                problems.append(f"{c['code']}: unknown category code '{c['type_code']}'")
            typ = ""
        cr = c["credits"]
        rows.append([c["code"], nm, cat, typ, 0 if cr == 0 else 3, cr,
                     c["L"] + c["T"], c["P"], 25, 75])
    return rows, problems


def main():
    args = sys.argv[1:]
    if not args or args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0 if args else 2)
    src = args[0]
    here = os.path.dirname(os.path.abspath(__file__))
    template = os.path.join(here, "..", "assets", "bos-courses-import-template.xlsx")
    out = os.path.splitext(src)[0] + "-import.xlsx"
    expected = None
    for i, a in enumerate(args):
        if a == "--template":
            template = args[i + 1]
        elif a == "--out":
            out = args[i + 1]
        elif a == "--expected-credits":
            expected = float(args[i + 1])

    import openpyxl
    courses, fmt = extract(src)
    print(f"Format detected: {fmt}")
    if not courses:
        print("No courses extracted — nothing written.")
        sys.exit(1)
    rows, problems = to_rows(courses)

    wb = openpyxl.load_workbook(template)
    ws = wb["Courses"]
    vcat, vtyp = load_valid_lists(wb)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    total = 0.0
    for r, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(r, col, val)
        if row[2] not in vcat:
            problems.append(f"{row[0]}: Category '{row[2]}' not in Lists")
        if row[3] and row[3] not in vtyp:
            problems.append(f"{row[0]}: Type '{row[3]}' not in Lists")
        total += float(row[5])
    wb.save(out)

    print(f"Wrote {out}")
    print(f"  Courses: {len(rows)}   Credits: {total:g}")
    if expected is not None:
        d = total - expected
        print(f"  {'Reconciles to ' + str(expected) if abs(d) < 1e-9 else f'Differs from {expected} by {d:g} (explain: NCC/extra-credit or elective pool)'}")
    if problems:
        print("\nVALIDATION ISSUES:")
        for p in problems:
            print("  -", p)
        sys.exit(1)
    print("\nAll Category/Type values valid against the Lists sheet.")


if __name__ == "__main__":
    main()
