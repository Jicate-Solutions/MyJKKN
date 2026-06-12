"""Generates Razorpay CollectNow audit "transaction details" workbooks
(Ticket #19383043) requested by CYRAACS, one per batch of order/payment IDs.

All values are REAL, pulled from production payment_transactions /
billing_student_bills. Timestamps converted UTC -> IST (+5:30).

Usage:  python scripts/gen-razorpay-transaction-xlsx.py [batch]
        batch = "1", "2", or "all" (default: all)
Output: C:/Users/Admin/Downloads/JKKN-Razorpay-Transaction-Details-Ticket-19383043[-batchN].xlsx
"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = "C:/Users/Admin/Downloads"

TEAL = "0F766E"
GREEN_BG = "E6F4EA"
RED_BG = "FCE8E6"
GREY = "555555"
HEAD_FILL = PatternFill("solid", fgColor=TEAL)
INFO_FILL = PatternFill("solid", fgColor="F0F4F8")
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
FONT = "Arial"

HEADERS = [
    "S.No", "Razorpay Order ID", "Razorpay Payment ID", "Transaction Status",
    "Transaction Amount (INR)", "Amount (paise)", "No. of DB Rows for Order ID",
    "Transaction Timestamp (IST)", "Captured / Completed (IST)",
    "Internal Txn Reference", "Receipt No.", "Product Name", "Product Type",
]
WIDTHS = [6, 22, 20, 18, 16, 13, 16, 24, 24, 22, 16, 40, 22]
NCOL = len(HEADERS)
LAST_COL = get_column_letter(NCOL)

CUSTOMER = "BOOBAL A — Roll No 87596328 · Reg No A123654789"

# ---- Real data, one entry per batch ----------------------------------------
# Row cols: S.No, OrderID, PaymentID, Status, Amount(INR), paise, DBrows,
#           Timestamp(IST), Captured(IST), InternalRef, Receipt, Product, Type
BATCHES = {
    "1": {
        "suffix": "-batch1",
        "window": "11-Jun-2026 17:11 to 17:39",
        "rows": [
            [1, "order_T0JMesiC9AAwk5", "pay_T0JNMHuscjv13z", "Success (captured)", 250.00, 25000, 1,
             "11-Jun-2026 17:11:47", "11-Jun-2026 17:12:27", "P202606111141472Y7TO", "RCP-2026-002321",
             "4 Year Tuition Fee - Razorpay Test Due 12", "Tuition Fee (Academic)"],
            [2, "order_T0JSBR6QkPPT8h", "pay_T0JSufYHUjUfZ1", "Failed", 25974.00, 2597400, 1,
             "11-Jun-2026 17:17:01", "—", "P20260611114701KVZ0F", "—",
             "3 Year Tuition Fee", "Tuition Fee (Academic)"],
            [3, "order_T0JXFiyMfTVWYR", "pay_T0JXN30ZTP2WGG", "Failed", 25974.00, 2597400, 1,
             "11-Jun-2026 17:21:49", "—", "P2026061111514988FBU", "—",
             "3 Year Tuition Fee", "Tuition Fee (Academic)"],
            [4, "order_T0JbKasUYbKtxz", "pay_T0Jby4cHJCgeJY", "Failed", 3000.00, 300000, 1,
             "11-Jun-2026 17:25:41", "—", "P20260611115540U4VAF", "—",
             "1 Year Tuition Fee - Razorpay Test Due 09", "Tuition Fee (Academic)"],
            [5, "order_T0JfsXn7SvYuB3", "pay_T0JgaHH2P3BUKL", "Failed", 3000.00, 300000, 1,
             "11-Jun-2026 17:29:59", "—", "P20260611115959VAXEU", "—",
             "1 Year Tuition Fee - Razorpay Test Due 09", "Tuition Fee (Academic)"],
            [6, "order_T0Jkq4Y8ZYPtkp", "—", "Failed", 3000.00, 300000, 1,
             "11-Jun-2026 17:34:41", "—", "P202606111204408GILP", "—",
             "1 Year Tuition Fee - Razorpay Test Due 09", "Tuition Fee (Academic)"],
            [7, "order_T0JosvZsSsK4ZL", "—", "Failed", 3000.00, 300000, 1,
             "11-Jun-2026 17:38:30", "—", "P202606111208305DIS4", "—",
             "1 Year Tuition Fee - Razorpay Test Due 09", "Tuition Fee (Academic)"],
        ],
    },
    "2": {
        "suffix": "-batch2",
        "window": "11-Jun-2026 17:49 to 12-Jun-2026 12:14",
        "rows": [
            [1, "order_T0K0oVr2QwO1nj", "pay_T0K16BOUKeWnSZ", "Success (captured)", 2500.00, 250000, 1,
             "11-Jun-2026 17:49:48", "11-Jun-2026 17:50:04", "P20260611121948G0A6U", "RCP-2026-002323",
             "Alumni Fee - Razorpay Test Due 08", "Alumni Fee (Other)"],
            [2, "order_T0K4ZSVlQjYowz", "pay_T0K4pD3w5rd7vs", "Failed", 3000.00, 300000, 1,
             "11-Jun-2026 17:53:21", "—", "P20260611122321879YD", "—",
             "1 Year Tuition Fee - Razorpay Test Due 09", "Tuition Fee (Academic)"],
            [3, "order_T0cGxEbXA3WlOn", "pay_T0cHTaTxW26a4Z", "Failed", 5000.00, 500000, 1,
             "12-Jun-2026 11:41:34", "—", "P20260612061134JPIX3", "—",
             "2 Year Tuition Fee - Razorpay Test Due 10", "Tuition Fee (Academic)"],
            [4, "order_T0cMgQU296FEP5", "pay_T0cN4cDH6XruYD", "Success (captured)", 2000.00, 200000, 1,
             "12-Jun-2026 11:46:59", "12-Jun-2026 11:47:22", "P20260612061659WSNES", "— (webhook-captured)",
             "Application Fee - Razorpay Test Due 07", "Application Fee"],
            [5, "order_T0cpdDHt9GO2tO", "pay_T0cqVC272c9Xf4", "Success (captured)", 2000.00, 200000, 1,
             "12-Jun-2026 12:14:23", "12-Jun-2026 12:15:13", "P20260612064423R4YH6", "RCP-2026-002332",
             "Application Fee - Razorpay Test Due 07", "Application Fee"],
        ],
    },
}


def build(batch_key):
    cfg = BATCHES[batch_key]
    rows = cfg["rows"]
    out = f"{OUT_DIR}/JKKN-Razorpay-Transaction-Details-Ticket-19383043{cfg['suffix']}.xlsx"

    wb = Workbook()
    # No local LibreOffice/Excel engine to pre-cache results, so force the
    # reader (Excel / Google Sheets) to evaluate the summary formulas on open.
    wb.calculation.fullCalcOnLoad = True
    ws = wb.active
    ws.title = "Transaction Details"
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def merge(rng):
        ws.merge_cells(rng)

    r = 1
    ws.cell(r, 1, "JKKN — Razorpay Transaction Details").font = Font(FONT, size=16, bold=True, color=TEAL)
    merge(f"A{r}:{LAST_COL}{r}"); ws.row_dimensions[r].height = 24; r += 1
    ws.cell(r, 1, "CollectNow Security Audit · Ticket #19383043 · prepared for CYRAACS").font = Font(FONT, size=10, color=GREY)
    merge(f"A{r}:{LAST_COL}{r}"); r += 2

    info = [
        ("Merchant / Application", "MyJKKN Platform — https://www.jkkn.ai"),
        ("Payment Provider", "Razorpay — Hosted Checkout (CollectNow)"),
        ("Institution", "JKKN Testing Institution"),
        ("Customer (Student)", CUSTOMER),
        ("Audit Window (IST)", cfg["window"]),
        ("Data Source", "Production DB: payment_transactions + billing_student_bills (as of 12-Jun-2026)"),
    ]
    for label, val in info:
        c = ws.cell(r, 1, label); c.font = Font(FONT, size=10, bold=True); c.fill = INFO_FILL
        c.border = BORDER; c.alignment = Alignment(vertical="center"); merge(f"A{r}:C{r}")
        v = ws.cell(r, 4, val); v.font = Font(FONT, size=10); v.border = BORDER
        v.alignment = Alignment(vertical="center"); merge(f"D{r}:{LAST_COL}{r}"); r += 1
    r += 1

    ws.cell(r, 1, "Summary").font = Font(FONT, size=12, bold=True, color=TEAL)
    merge(f"A{r}:{LAST_COL}{r}"); r += 1
    summary_start = r
    labels = [
        "Total transactions queried", "Successful (captured)", "Failed",
        "Max DB rows for any single Order ID (duplicate check)", "Distinct Order IDs",
        "Total successful amount (INR)",
    ]
    for label in labels:
        c = ws.cell(r, 1, label); c.font = Font(FONT, size=10, bold=True); c.fill = INFO_FILL
        c.border = BORDER; merge(f"A{r}:E{r}"); ws.cell(r, 6).border = BORDER; r += 1
    r += 1

    header_row = r
    for ci, h in enumerate(HEADERS, start=1):
        c = ws.cell(header_row, ci, h)
        c.font = Font(FONT, size=10, bold=True, color="FFFFFF"); c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = BORDER
    ws.row_dimensions[header_row].height = 30

    data_start = header_row + 1
    for ri, row in enumerate(rows):
        rr = data_start + ri
        is_success = "Success" in row[3]
        fill = PatternFill("solid", fgColor=GREEN_BG if is_success else RED_BG)
        for ci, val in enumerate(row, start=1):
            c = ws.cell(rr, ci, val); c.font = Font(FONT, size=10); c.border = BORDER; c.fill = fill
            if ci in (1, 4, 6, 7):
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 5:
                c.number_format = '"₹"#,##0.00'; c.alignment = Alignment(horizontal="right", vertical="center")
            elif ci == 6:
                c.number_format = '#,##0'
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(ci == 12))
            if ci == 4:
                c.font = Font(FONT, size=10, bold=True, color=("0F766E" if is_success else "B3261E"))
    data_end = data_start + len(rows) - 1

    D = f"D{data_start}:D{data_end}"; E = f"E{data_start}:E{data_end}"
    G = f"G{data_start}:G{data_end}"; B = f"B{data_start}:B{data_end}"
    formulas = [
        f"=COUNTA({B})", f'=COUNTIF({D},"*Success*")', f'=COUNTIF({D},"Failed")',
        f"=MAX({G})", f"=SUMPRODUCT(1/COUNTIF({B},{B}))", f'=SUMIF({D},"*Success*",{E})',
    ]
    for i, f in enumerate(formulas):
        cell = ws.cell(summary_start + i, 6, f); cell.font = Font(FONT, size=10, bold=True); cell.border = BORDER
        if i == 5:
            cell.number_format = '"₹"#,##0.00'

    r = data_end + 2
    note = ("Note: 'No. of DB Rows for Order ID' = number of rows persisted in payment_transactions for that "
            "Razorpay order_id. A partial UNIQUE index (razorpay_order_id) makes a value of 1 the structurally "
            "enforced maximum — each order ID is stored exactly once. Failed/abandoned attempts are persisted "
            "as required by the audit checklist. Successful payments are confirmed server-side (HMAC signature "
            "+ dual inquiry: GET /v1/orders + GET /v1/payments, amount matched to the paise) before the database "
            "is marked 'success' and the receipt/credit is granted.")
    c = ws.cell(r, 1, note); c.font = Font(FONT, size=9, italic=True, color=GREY)
    c.alignment = Alignment(wrap_text=True, vertical="top"); merge(f"A{r}:{LAST_COL}{r}")
    ws.row_dimensions[r].height = 80

    wb.save(out)
    print("WROTE", out)


arg = sys.argv[1] if len(sys.argv) > 1 else "all"
keys = list(BATCHES) if arg == "all" else [arg]
for k in keys:
    build(k)
